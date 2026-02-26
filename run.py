"""
=============================================================
  LEADSCRAPER — Backend API v2
  FastAPI + WebSocket + SQLite
=============================================================
  ARRANQUE: py -3.11 run.py
=============================================================
"""

import sys
import asyncio
import concurrent.futures
import json
import os
import sqlite3
import time
import traceback
import uuid
from datetime import datetime
from typing import Optional

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel

# ─────────────────────────────────────────────
#  APP
# ─────────────────────────────────────────────
app = FastAPI(title="LeadScraper API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DATA_DIR  = os.path.join(BASE_DIR, "..", "data")
DB_PATH   = os.path.join(DATA_DIR, "leadscraper.db")
EXCEL_DIR = os.path.join(DATA_DIR, "exports")

os.makedirs(DATA_DIR,  exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

# ─────────────────────────────────────────────
#  BASE DE DADOS
# ─────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS sessoes (
            id          TEXT PRIMARY KEY,
            criado_em   TEXT NOT NULL,
            sector      TEXT,
            zonas       TEXT,
            filtros     TEXT,
            estado      TEXT DEFAULT 'em_curso',
            total_leads INTEGER DEFAULT 0,
            duracao_seg INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS leads (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            sessao_id   TEXT NOT NULL,
            nome        TEXT,
            telefone    TEXT,
            website     TEXT,
            morada      TEXT,
            avaliacoes  TEXT,
            qualidade   TEXT,
            ramo        TEXT,
            pesquisa    TEXT,
            url_maps    TEXT,
            estado      TEXT DEFAULT 'Nao contactado',
            notas       TEXT DEFAULT '',
            criado_em   TEXT NOT NULL,
            FOREIGN KEY (sessao_id) REFERENCES sessoes(id)
        );
        CREATE INDEX IF NOT EXISTS idx_leads_sessao ON leads(sessao_id);
        CREATE INDEX IF NOT EXISTS idx_leads_qualidade ON leads(qualidade);
    """)
    conn.commit()
    conn.close()

init_db()

# ─────────────────────────────────────────────
#  ESTADO GLOBAL
# ─────────────────────────────────────────────
sessoes_activas: dict = {}
thread_pool = concurrent.futures.ThreadPoolExecutor(max_workers=3)

class ConnectionManager:
    def __init__(self):
        self.connections: dict = {}

    async def connect(self, sessao_id: str, ws: WebSocket):
        await ws.accept()
        self.connections.setdefault(sessao_id, []).append(ws)

    def disconnect(self, sessao_id: str, ws: WebSocket):
        if sessao_id in self.connections:
            try:
                self.connections[sessao_id].remove(ws)
            except ValueError:
                pass

    async def broadcast(self, sessao_id: str, msg: dict):
        for ws in self.connections.get(sessao_id, []):
            try:
                await ws.send_json(msg)
            except Exception:
                pass

manager = ConnectionManager()

# ─────────────────────────────────────────────
#  MODELOS
# ─────────────────────────────────────────────
class ConfiguracaoPesquisa(BaseModel):
    sector:          str
    zonas:           list
    tipos:           list
    so_com_telefone: bool = True
    so_sem_website:  bool = False
    max_resultados:  int  = 15
    sheets_url:      Optional[str] = None

class ComandoSessao(BaseModel):
    comando: str

# ─────────────────────────────────────────────
#  WEBSOCKET — envio thread-safe
# ─────────────────────────────────────────────
_main_loop: Optional[asyncio.AbstractEventLoop] = None

@app.on_event("startup")
async def startup():
    global _main_loop
    _main_loop = asyncio.get_event_loop()

def enviar_sync(sessao_id: str, msg: dict):
    """Envia mensagem WebSocket a partir de uma thread (thread-safe)."""
    if _main_loop and _main_loop.is_running():
        asyncio.run_coroutine_threadsafe(
            manager.broadcast(sessao_id, msg),
            _main_loop
        )

# ─────────────────────────────────────────────
#  SCRAPER (corre em thread separada)
# ─────────────────────────────────────────────
def scrape_em_thread(sessao_id: str, config: ConfiguracaoPesquisa):
    """Corre numa thread com ProactorEventLoop próprio."""
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_scraper(sessao_id, config))
    except Exception as e:
        enviar_sync(sessao_id, {"tipo": "erro", "mensagem": str(e)[:200]})
    finally:
        loop.close()

async def _scraper(sessao_id: str, config: ConfiguracaoPesquisa):
    from playwright.async_api import async_playwright

    sessao     = sessoes_activas[sessao_id]
    inicio     = time.time()
    total_leads = 0

    def env(tipo, dados={}):
        enviar_sync(sessao_id, {"tipo": tipo, "sessao_id": sessao_id, **dados})

    env("inicio", {"mensagem": "Scraper iniciado!"})

    pesquisas = [f"{tipo} {zona}" for tipo in config.tipos for zona in config.zonas]

    try:
        async with async_playwright() as p:
            # Em servidor (Linux) usa headless. Em Windows local usa visível
            is_server = sys.platform != "win32"

            browser = await p.chromium.launch(
                headless=is_server,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-gpu",
                    "--window-size=1280,800",
                ]
            )
            context = await browser.new_context(
                locale="pt-PT",
                viewport={"width": 1280, "height": 800},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
            )
            page = await context.new_page()

            # Parecer humano
            await page.goto("https://www.google.pt", wait_until="domcontentloaded")
            await page.wait_for_timeout(2000)

            for idx, pesquisa in enumerate(pesquisas, 1):

                # Verificar comandos
                while sessao["estado"] == "pausado":
                    env("pausado", {"mensagem": "Pausado..."})
                    await asyncio.sleep(2)

                if sessao["estado"] == "parado":
                    break

                env("pesquisa", {"pesquisa": pesquisa, "idx": idx, "total": len(pesquisas)})

                url = f"https://www.google.com/maps/search/{pesquisa.replace(' ', '+')}?hl=pt-PT"
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    await page.wait_for_timeout(3500)
                except Exception as e:
                    env("aviso", {"mensagem": f"Erro ao carregar: {str(e)[:60]}"}); continue

                # Rolar resultados
                rolou = False
                for sel in ['[role="feed"]', 'div[aria-label*="Resultados"]', 'div[aria-label*="Results"]']:
                    try:
                        await page.wait_for_selector(sel, timeout=6000)
                        for _ in range(5):
                            await page.evaluate(f"document.querySelector('{sel}').scrollTop += 2000")
                            await page.wait_for_timeout(900)
                        rolou = True
                        break
                    except Exception:
                        continue
                if not rolou:
                    for _ in range(5):
                        await page.keyboard.press("End")
                        await page.wait_for_timeout(700)

                # Recolher links
                links = await page.evaluate("""
                    () => {
                        const s = new Set();
                        document.querySelectorAll('a[href*="/maps/place/"]').forEach(el => {
                            const h = el.href || el.getAttribute('href') || '';
                            if (h.includes('/maps/place/'))
                                s.add(h.startsWith('http') ? h : 'https://www.google.com' + h);
                        });
                        return Array.from(s);
                    }
                """)
                links = list(set(links))[:config.max_resultados]
                env("aviso", {"mensagem": f"{pesquisa}: {len(links)} resultados"})

                for link in links:
                    if sessao["estado"] == "parado":
                        break

                    try:
                        await page.goto(link, wait_until="domcontentloaded", timeout=20000)
                        await page.wait_for_timeout(1800)

                        nome = await page.evaluate("() => { const h = document.querySelector('h1'); return h ? h.innerText.trim() : ''; }")
                        if not nome:
                            continue

                        telefone = await page.evaluate("""
                            () => {
                                for (const b of document.querySelectorAll('button[data-item-id^="phone"]')) {
                                    const m = (b.getAttribute('data-item-id')||'').match(/phone:tel:(.+)/);
                                    if (m) return m[1];
                                }
                                const m = document.body.innerText.match(/(\+351[\s]?)?[29][0-9]{8}/);
                                return m ? m[0].trim() : '';
                            }
                        """)

                        if config.so_com_telefone and not telefone:
                            continue

                        website = await page.evaluate("""
                            () => {
                                for (const el of document.querySelectorAll('a[data-item-id="authority"]')) {
                                    const h = el.getAttribute('href') || '';
                                    if (h && !h.includes('google.com')) return h;
                                }
                                return '';
                            }
                        """)

                        tem_website = bool(website and len(website) > 5)
                        if config.so_sem_website and tem_website:
                            continue

                        morada = await page.evaluate("""
                            () => {
                                for (const b of document.querySelectorAll('button[data-item-id="address"]'))
                                    return b.innerText.trim();
                                return '';
                            }
                        """)

                        avaliacoes = await page.evaluate("""
                            () => {
                                const m = document.body.innerText.match(/\((\d+)\s*(avalia|review)/i);
                                return m ? m[1] : '0';
                            }
                        """)

                        if not tem_website:
                            qualidade = "QUENTE"
                        elif any(x in website.lower() for x in ['wix','blogspot','wordpress.com','jimdo']):
                            qualidade = "MEDIO"
                        else:
                            qualidade = "FRIO"

                        lead = {
                            "nome":       nome,
                            "telefone":   telefone,
                            "website":    website if tem_website else "SEM WEBSITE",
                            "morada":     morada,
                            "avaliacoes": avaliacoes,
                            "qualidade":  qualidade,
                            "ramo":       config.sector,
                            "pesquisa":   pesquisa,
                            "url_maps":   link,
                        }

                        # Guardar na BD
                        conn = get_db()
                        conn.execute("""
                            INSERT INTO leads
                            (sessao_id,nome,telefone,website,morada,avaliacoes,
                             qualidade,ramo,pesquisa,url_maps,criado_em)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        """, (sessao_id, nome, telefone,
                              website if tem_website else "SEM WEBSITE",
                              morada, avaliacoes, qualidade,
                              config.sector, pesquisa, link,
                              datetime.now().isoformat()))
                        conn.commit()
                        total_leads += 1
                        conn.execute("UPDATE sessoes SET total_leads=? WHERE id=?",
                                     (total_leads, sessao_id))
                        conn.commit()
                        conn.close()

                        env("lead", {"lead": lead, "total": total_leads})

                    except Exception as e:
                        env("aviso", {"mensagem": f"Erro num lead: {str(e)[:60]}"}); continue

                await asyncio.sleep(1.5)

            await browser.close()

    except Exception as e:
        env("erro", {"mensagem": traceback.format_exc()[:300]})

    duracao = int(time.time() - inicio)
    conn = get_db()
    conn.execute("UPDATE sessoes SET estado='concluido', total_leads=?, duracao_seg=? WHERE id=?",
                 (total_leads, duracao, sessao_id))
    conn.commit()
    conn.close()

    sessao["estado"] = "concluido"
    env("concluido", {
        "total":    total_leads,
        "duracao":  duracao,
        "mensagem": f"Concluído! {total_leads} leads em {duracao//60}m{duracao%60}s"
    })

# ─────────────────────────────────────────────
#  ENDPOINTS REST
# ─────────────────────────────────────────────
@app.post("/api/sessao/iniciar")
async def iniciar_sessao(config: ConfiguracaoPesquisa):
    sessao_id = str(uuid.uuid4())[:8]

    conn = get_db()
    conn.execute("""
        INSERT INTO sessoes (id, criado_em, sector, zonas, filtros, estado)
        VALUES (?, ?, ?, ?, ?, 'em_curso')
    """, (sessao_id, datetime.now().isoformat(), config.sector,
          json.dumps(config.zonas),
          json.dumps({"so_com_telefone": config.so_com_telefone,
                      "so_sem_website":  config.so_sem_website})))
    conn.commit()
    conn.close()

    sessoes_activas[sessao_id] = {"estado": "correndo", "config": config}
    thread_pool.submit(scrape_em_thread, sessao_id, config)

    return {"sessao_id": sessao_id, "mensagem": "Sessão iniciada!"}


@app.post("/api/sessao/{sessao_id}/comando")
async def comando_sessao(sessao_id: str, cmd: ComandoSessao):
    if sessao_id not in sessoes_activas:
        raise HTTPException(404, "Sessão não encontrada")
    sessao = sessoes_activas[sessao_id]
    if cmd.comando == "pausar":
        sessao["estado"] = "pausado"
    elif cmd.comando == "continuar":
        sessao["estado"] = "correndo"
    elif cmd.comando == "parar":
        sessao["estado"] = "parado"
        conn = get_db()
        conn.execute("UPDATE sessoes SET estado='parado' WHERE id=?", (sessao_id,))
        conn.commit()
        conn.close()
    else:
        raise HTTPException(400, f"Comando inválido: {cmd.comando}")
    return {"ok": True, "estado": sessao["estado"]}


@app.get("/api/sessao/{sessao_id}/leads")
async def listar_leads(sessao_id: str, qualidade: str = "", pagina: int = 1, por_pagina: int = 50):
    conn   = get_db()
    where  = "WHERE sessao_id = ?"
    params = [sessao_id]
    if qualidade:
        where += " AND qualidade = ?"
        params.append(qualidade)
    total  = conn.execute(f"SELECT COUNT(*) FROM leads {where}", params).fetchone()[0]
    offset = (pagina - 1) * por_pagina
    rows   = conn.execute(
        f"SELECT * FROM leads {where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [por_pagina, offset]
    ).fetchall()
    conn.close()
    return {"total": total, "pagina": pagina, "leads": [dict(r) for r in rows]}


@app.get("/api/historico")
async def historico():
    conn = get_db()
    rows = conn.execute("""
        SELECT id, criado_em, sector, zonas, estado, total_leads, duracao_seg
        FROM sessoes ORDER BY criado_em DESC LIMIT 50
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/sessao/{sessao_id}/stats")
async def stats_sessao(sessao_id: str):
    conn    = get_db()
    total   = conn.execute("SELECT COUNT(*) FROM leads WHERE sessao_id=?", (sessao_id,)).fetchone()[0]
    quentes = conn.execute("SELECT COUNT(*) FROM leads WHERE sessao_id=? AND qualidade='QUENTE'", (sessao_id,)).fetchone()[0]
    medios  = conn.execute("SELECT COUNT(*) FROM leads WHERE sessao_id=? AND qualidade='MEDIO'",  (sessao_id,)).fetchone()[0]
    frios   = conn.execute("SELECT COUNT(*) FROM leads WHERE sessao_id=? AND qualidade='FRIO'",   (sessao_id,)).fetchone()[0]
    conn.close()
    return {"total": total, "quentes": quentes, "medios": medios, "frios": frios}


@app.get("/api/sessao/{sessao_id}/exportar/excel")
async def exportar_excel(sessao_id: str):
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment

    conn = get_db()
    rows = conn.execute("SELECT * FROM leads WHERE sessao_id=?", (sessao_id,)).fetchall()
    sessao_row = conn.execute("SELECT * FROM sessoes WHERE id=?", (sessao_id,)).fetchone()
    conn.close()

    if not rows:
        raise HTTPException(404, "Nenhum lead encontrado")

    df   = pd.DataFrame([dict(r) for r in rows])
    df   = df.drop(columns=["id", "sessao_id"], errors="ignore")
    path = os.path.join(EXCEL_DIR, f"leads_{sessao_id}.xlsx")
    cores = {"QUENTE": "FF6B6B", "MEDIO": "FFB347", "FRIO": "90EE90"}

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        ws = writer.sheets["Leads"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cor  = cores.get(row[6].value, "FFFFFF")
            fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
            for cell in row:
                cell.fill = fill
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes      = "A2"
        ws.auto_filter.ref   = ws.dimensions

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"leads_{sessao_row['sector']}_{sessao_id}.xlsx"
    )


# ─────────────────────────────────────────────
#  WEBSOCKET
# ─────────────────────────────────────────────
@app.websocket("/ws/{sessao_id}")
async def websocket_endpoint(ws: WebSocket, sessao_id: str):
    await manager.connect(sessao_id, ws)
    try:
        while True:
            await asyncio.sleep(20)
            await ws.send_json({"tipo": "ping"})
    except WebSocketDisconnect:
        manager.disconnect(sessao_id, ws)
    except Exception:
        manager.disconnect(sessao_id, ws)


# ─────────────────────────────────────────────
#  FRONTEND
# ─────────────────────────────────────────────
FRONTEND_DIR = os.path.join(BASE_DIR, "..", "frontend")
if os.path.exists(FRONTEND_DIR):
    app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")

    @app.get("/")
    async def root():
        return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))
